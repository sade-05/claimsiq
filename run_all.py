# =============================================================================
# ClaimsIQ — No-Fault Claims Intelligence System
# run_all.py  |  Master pipeline — all three phases in one file
# =============================================================================
# Usage:
#   python run_all.py
#
# Phases:
#   Phase 1 — Ingest, Clean & Store
#   Phase 2 — Forecast, Score & Map
#   Phase 3 — Report & Deliver
#
# Stops immediately if any phase fails.
# =============================================================================

import pandas as pd
import numpy as np
import sqlite3
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import os
import datetime
import warnings
import sys
import time
warnings.filterwarnings("ignore")

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable, Image, PageBreak
)
from reportlab.lib.colors import HexColor

# ── Config ────────────────────────────────────────────────────────────────────
DATA_FILE = "data/insurance_claims.csv"
DB_FILE   = "outputs/claims.db"
OUT_DIR   = "outputs"

# =============================================================================
# PHASE 1 — INGEST, CLEAN & STORE
# =============================================================================

COLUMN_MAP = {
    "policy_number":              "Policy Number",
    "policy_bind_date":           "Policy Bind Date",
    "policy_state":               "Policy State",
    "policy_csl":                 "Policy CSL",
    "policy_deductable":          "Deductible ($)",
    "policy_annual_premium":      "Annual Premium ($)",
    "umbrella_limit":             "Umbrella Limit",
    "insured_zip":                "Claimant ZIP",
    "insured_sex":                "Claimant Gender",
    "insured_education_level":    "Education Level",
    "insured_occupation":         "Occupation",
    "insured_hobbies":            "Claimant Background",
    "insured_relationship":       "Relationship to Insured",
    "capital-gains":              "Capital Gains",
    "capital-loss":               "Capital Loss",
    "incident_date":              "Date of Loss",
    "incident_type":              "Accident Type",
    "collision_type":             "Collision Type",
    "incident_severity":          "Incident Severity",
    "authorities_contacted":      "Authorities Notified",
    "incident_state":             "State",
    "incident_city":              "City",
    "incident_location":          "Incident Location",
    "incident_hour_of_the_day":   "Hour of Incident",
    "number_of_vehicles_involved":"Vehicles in Accident",
    "property_damage":            "Property Damage",
    "bodily_injuries":            "Reported Injuries",
    "witnesses":                  "Witness Count",
    "police_report_available":    "Police Report on File",
    "total_claim_amount":         "Total Amount Billed ($)",
    "injury_claim":               "Medical Bills ($)",
    "property_claim":             "Property Damage ($)",
    "vehicle_claim":              "Vehicle Damage ($)",
    "auto_make":                  "Vehicle Make",
    "auto_model":                 "Vehicle Model",
    "auto_year":                  "Vehicle Year",
    "fraud_reported":             "Fraud Flag",
    "months_as_customer":         "Policy Tenure (Months)",
    "age":                        "Claimant Age",
}

def phase1():
    print("\n" + "="*60)
    print("  PHASE 1 — INGEST, CLEAN & STORE")
    print("="*60)

    if not os.path.exists(DATA_FILE):
        print(f"\n  ERROR: '{DATA_FILE}' not found.")
        print("  Please download insurance_claims.csv from Kaggle and")
        print("  place it in the data/ folder, then run again.\n")
        print("  Download: kaggle.com/datasets/buntyshah/auto-insurance-claims-data\n")
        sys.exit(1)

    df = pd.read_csv(DATA_FILE)
    print(f"\n  Loaded {len(df):,} claims from '{DATA_FILE}'")

    df.replace("?", pd.NA, inplace=True)
    df.rename(columns={k: v for k, v in COLUMN_MAP.items() if k in df.columns}, inplace=True)
    print(f"  Remapped {len(COLUMN_MAP)} columns to no-fault labels")

    if "Date of Loss" in df.columns:
        df["Date of Loss"] = pd.to_datetime(df["Date of Loss"], errors="coerce")
        df["Week"]  = df["Date of Loss"].dt.to_period("W").astype(str)
        df["Month"] = df["Date of Loss"].dt.to_period("M").astype(str)
        df["Year"]  = df["Date of Loss"].dt.year

    if "Fraud Flag" in df.columns:
        df["Fraud Flag"] = df["Fraud Flag"].str.strip().str.upper()

    for col in ["Total Amount Billed ($)", "Medical Bills ($)",
                "Property Damage ($)", "Vehicle Damage ($)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    os.makedirs(OUT_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    df.to_sql("claims", conn, if_exists="replace", index=False)
    conn.close()
    print(f"  Written to database: {DB_FILE}")
    print(f"  Table 'claims' — {len(df):,} rows, {len(df.columns)} columns")
    print("\n  Phase 1 complete.\n")

# =============================================================================
# PHASE 2 — FORECAST, SCORE & MAP
# =============================================================================

BLUE   = "#2E75B6"
ORANGE = "#ED7D31"
RED    = "#C00000"
GREEN  = "#70AD47"
GRAY   = "#AAAAAA"

plt.rcParams.update({
    "font.family":       "DejaVu Sans",
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.grid":         True,
    "grid.alpha":        0.3,
    "grid.linestyle":    "--",
    "figure.dpi":        150,
})

RISK_COLORS = {"LOW": GREEN, "MEDIUM": "#FFD966", "HIGH": ORANGE, "CRITICAL": RED}

def _load_claims():
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql("SELECT * FROM claims", conn)
    conn.close()
    df["Date of Loss"] = pd.to_datetime(df["Date of Loss"], errors="coerce")
    return df

def _save_chart(fig, name):
    path = os.path.join(OUT_DIR, name)
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Saved: {path}")

def simple_arima(series, steps=8):
    try:
        from pmdarima import auto_arima
        model = auto_arima(series, seasonal=False, suppress_warnings=True, error_action="ignore")
        fc, ci = model.predict(n_periods=steps, return_conf_int=True)
        return fc, ci[:, 0], ci[:, 1]
    except ImportError:
        pass
    try:
        from statsmodels.tsa.holtwinters import ExponentialSmoothing
        model = ExponentialSmoothing(series, trend="add").fit()
        fc = model.forecast(steps)
        std = series.std()
        return fc.values, fc.values - 1.5 * std, fc.values + 1.5 * std
    except ImportError:
        pass
    x = np.arange(len(series))
    m, b = np.polyfit(x, series.values, 1)
    fc = np.array([m * (len(series) + i) + b for i in range(steps)])
    std = series.std()
    return fc, fc - 1.5 * std, fc + 1.5 * std

def score_claim(row):
    score = 0
    flags = []
    try:
        if float(row.get("Vehicles in Accident", 1)) >= 4:
            score += 30
            flags.append(f"High claimants ({int(row['Vehicles in Accident'])})")
    except (ValueError, TypeError):
        pass
    if str(row.get("Fraud Flag", "N")).strip().upper() == "Y":
        score += 25
        flags.append("Fraud flagged")
    try:
        billed = float(row.get("Total Amount Billed ($)", 0))
        if billed > 50000:
            score += 25
            flags.append(f"High bill (${billed:,.0f})")
        elif billed > 25000:
            score += 15
            flags.append(f"Elevated bill (${billed:,.0f})")
    except (ValueError, TypeError):
        pass
    if str(row.get("Police Report on File", "YES")).strip().upper() == "NO":
        score += 10
        flags.append("No police report")
    try:
        if int(row.get("Witness Count", 1)) == 0:
            score += 10
            flags.append("No witnesses")
    except (ValueError, TypeError):
        pass
    if str(row.get("Incident Severity", "")).strip().lower() in ["total loss", "major damage"]:
        score += 10
        flags.append("Severe damage")
    return score, "; ".join(flags) if flags else "None"

def risk_label(score):
    if score >= 75: return "CRITICAL"
    if score >= 46: return "HIGH"
    if score >= 21: return "MEDIUM"
    return "LOW"

def chart_weekly_volume(df):
    weekly = df.groupby("Week").size().reset_index(name="Claims")
    weekly = weekly.sort_values("Week").reset_index(drop=True)
    weekly["MA4"] = weekly["Claims"].rolling(4, min_periods=1).mean()
    fig, ax = plt.subplots(figsize=(11, 4))
    ax.bar(range(len(weekly)), weekly["Claims"], color=BLUE, alpha=0.45, label="Weekly claims")
    ax.plot(range(len(weekly)), weekly["MA4"], color=ORANGE, lw=2.2, label="4-week moving average")
    step = max(1, len(weekly) // 8)
    ax.set_xticks(range(0, len(weekly), step))
    ax.set_xticklabels(weekly["Week"].iloc[::step], rotation=35, ha="right", fontsize=8)
    ax.set_xlabel("Week")
    ax.set_ylabel("Number of claims")
    ax.set_title("Weekly claim volume with 4-week moving average", fontweight="bold", pad=10)
    ax.legend()
    fig.tight_layout()
    _save_chart(fig, "chart_1_weekly_volume.png")
    return weekly

def chart_arima_forecast(weekly):
    series = weekly["Claims"].copy()
    fc, lo, hi = simple_arima(series, steps=8)
    future_idx = list(range(len(series), len(series) + 8))
    fig, ax = plt.subplots(figsize=(11, 4))
    ax.bar(range(len(series)), series, color=BLUE, alpha=0.35, label="Actual claims")
    ax.plot(range(len(series)), weekly["MA4"], color=ORANGE, lw=1.8, label="Moving average")
    ax.plot(future_idx, fc, color=RED, lw=2.2, linestyle="--", label="ARIMA forecast")
    ax.fill_between(future_idx, lo, hi, color=RED, alpha=0.12, label="Confidence band")
    ax.axvline(len(series) - 0.5, color=GRAY, lw=1, linestyle=":")
    step = max(1, len(weekly) // 8)
    all_len = len(series) + 8
    ax.set_xticks(range(0, all_len, max(step, 1)))
    tick_labels = list(weekly["Week"].iloc[::step]) + [""] * 8
    ax.set_xticklabels(tick_labels[:len(range(0, all_len, max(step, 1)))],
                       rotation=35, ha="right", fontsize=8)
    ax.set_xlabel("Week")
    ax.set_ylabel("Number of claims")
    ax.set_title("Moving average vs ARIMA forecast — next 8 weeks", fontweight="bold", pad=10)
    ax.legend()
    fig.tight_layout()
    _save_chart(fig, "chart_2_ma_vs_arima.png")
    print(f"\n  ARIMA 8-week forecast:")
    for i, (f, l, h) in enumerate(zip(fc, lo, hi), 1):
        print(f"    Week +{i}: {f:.0f} claims  (range: {max(0,l):.0f} – {h:.0f})")

def chart_seasonality(df):
    df2 = df.copy()
    df2["Month Name"] = df2["Date of Loss"].dt.strftime("%b")
    df2["Month Num"]  = df2["Date of Loss"].dt.month
    monthly = (df2.groupby(["Month Num", "Month Name"])
                  .size().reset_index(name="Claims")
                  .sort_values("Month Num"))
    fig, ax = plt.subplots(figsize=(10, 4))
    bars = ax.bar(monthly["Month Name"], monthly["Claims"],
                  color=BLUE, alpha=0.75, edgecolor="white")
    for bar in bars:
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 1, str(int(bar.get_height())),
                ha="center", va="bottom", fontsize=9)
    ax.set_xlabel("Month")
    ax.set_ylabel("Total claims")
    ax.set_title("Seasonal claim distribution by month", fontweight="bold", pad=10)
    fig.tight_layout()
    _save_chart(fig, "chart_3_seasonality.png")

def chart_fraud_volume(df):
    weekly = df.groupby("Week").agg(
        Total=("Fraud Flag", "count"),
        Fraud=("Fraud Flag", lambda x: (x == "Y").sum())
    ).reset_index().sort_values("Week")
    weekly["Fraud Rate %"] = (weekly["Fraud"] / weekly["Total"] * 100).round(1)
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 6), sharex=True)
    ax1.bar(range(len(weekly)), weekly["Total"], color=BLUE, alpha=0.5, label="All claims")
    ax1.bar(range(len(weekly)), weekly["Fraud"], color=RED, alpha=0.7, label="Fraud flagged")
    ax1.set_ylabel("Claims")
    ax1.set_title("Claim volume vs fraud-flagged claims over time", fontweight="bold", pad=10)
    ax1.legend()
    ax2.plot(range(len(weekly)), weekly["Fraud Rate %"], color=RED, lw=2)
    ax2.fill_between(range(len(weekly)), weekly["Fraud Rate %"], alpha=0.15, color=RED)
    ax2.set_ylabel("Fraud rate (%)")
    ax2.set_xlabel("Week")
    step = max(1, len(weekly) // 8)
    ax2.set_xticks(range(0, len(weekly), step))
    ax2.set_xticklabels(weekly["Week"].iloc[::step], rotation=35, ha="right", fontsize=8)
    fig.tight_layout()
    _save_chart(fig, "chart_4_fraud_vs_volume.png")

def chart_top_cities(df):
    if "City" not in df.columns:
        print("  Skipping chart 5 — City column not found.")
        return
    city_totals = (df.groupby("City")["Total Amount Billed ($)"]
                     .sum().nlargest(10).sort_values().reset_index())
    city_totals.columns = ["City", "Total Billed ($)"]
    avg_billed = city_totals["Total Billed ($)"].mean()
    bar_colors = [RED if v > avg_billed else BLUE for v in city_totals["Total Billed ($)"]]
    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.barh(city_totals["City"], city_totals["Total Billed ($)"] / 1000,
                   color=bar_colors, alpha=0.82, edgecolor="white")
    for bar in bars:
        ax.text(bar.get_width() + 0.5,
                bar.get_y() + bar.get_height() / 2,
                f"${bar.get_width():.0f}K", va="center", fontsize=9)
    ax.set_xlabel("Total amount billed ($000s)")
    ax.set_title("Top 10 cities by total amount billed", fontweight="bold", pad=10)
    ax.legend(handles=[
        Patch(color=RED,  alpha=0.82, label="Above average — review"),
        Patch(color=BLUE, alpha=0.82, label="Below average"),
    ])
    fig.tight_layout()
    _save_chart(fig, "chart_5_top_cities.png")

def chart_risk_distribution(df):
    counts = df["Risk Level"].value_counts()
    order  = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]
    counts = counts.reindex([x for x in order if x in counts.index])
    bar_colors = [RISK_COLORS[r] for r in counts.index]
    fig, ax = plt.subplots(figsize=(7, 4))
    bars = ax.bar(counts.index, counts.values, color=bar_colors, alpha=0.85, edgecolor="white")
    for bar, val in zip(bars, counts.values):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.5, str(val),
                ha="center", va="bottom", fontweight="bold")
    ax.set_xlabel("Risk level")
    ax.set_ylabel("Number of claims")
    ax.set_title("Fraud risk level distribution across all claims", fontweight="bold", pad=10)
    fig.tight_layout()
    _save_chart(fig, "chart_6_risk_distribution.png")

def reserve_forecast(df):
    monthly = df.groupby("Month")["Total Amount Billed ($)"].sum().reset_index()
    monthly = monthly.sort_values("Month").reset_index(drop=True)
    if len(monthly) < 3:
        print("  Not enough monthly data for reserve forecast.")
        return
    fc, lo, hi = simple_arima(monthly["Total Amount Billed ($)"], steps=3)
    print(f"\n  Loss reserve forecast (next 3 months):")
    for label, f, l, h in zip(["30-day", "60-day", "90-day"], fc, lo, hi):
        print(f"    {label}: ${max(0,f):>10,.0f}  (range: ${max(0,l):,.0f} – ${max(0,h):,.0f})")

def city_alerts(df):
    if "City" not in df.columns:
        return
    city_weekly = (df.groupby(["City", "Week"])["Total Amount Billed ($)"]
                     .sum().reset_index().sort_values("Week"))
    alerts = []
    for city, grp in city_weekly.groupby("City"):
        if len(grp) < 3:
            continue
        grp = grp.sort_values("Week").reset_index(drop=True)
        rolling_avg = grp["Total Amount Billed ($)"].rolling(4, min_periods=2).mean()
        last_actual = grp["Total Amount Billed ($)"].iloc[-1]
        last_avg    = rolling_avg.iloc[-1]
        if last_avg > 0 and last_actual > last_avg * 1.3:
            pct = (last_actual - last_avg) / last_avg * 100
            alerts.append((city, last_actual, last_avg, pct))
    if alerts:
        print(f"\n  City billing alerts ({len(alerts)} flagged — > 30% above rolling avg):")
        for city, actual, avg, pct in sorted(alerts, key=lambda x: -x[3]):
            print(f"    {city:<20} actual: ${actual:>8,.0f}  avg: ${avg:>8,.0f}  (+{pct:.0f}%)")
    else:
        print("\n  No city billing alerts — all within normal range.")

def phase2():
    print("\n" + "="*60)
    print("  PHASE 2 — FORECAST, SCORE & MAP")
    print("="*60)

    if not os.path.exists(DB_FILE):
        print("  ERROR: claims.db not found. Phase 1 may have failed.")
        sys.exit(1)

    os.makedirs(OUT_DIR, exist_ok=True)
    df = _load_claims()
    print(f"\n  Loaded {len(df):,} claims from database")

    print("\n  Scoring fraud risk...")
    results = df.apply(lambda r: score_claim(r.to_dict()), axis=1)
    df["Risk Score"] = [r[0] for r in results]
    df["Risk Flags"] = [r[1] for r in results]
    df["Risk Level"] = df["Risk Score"].apply(risk_label)

    conn = sqlite3.connect(DB_FILE)
    df.to_sql("claims_scored", conn, if_exists="replace", index=False)
    conn.close()

    csv_cols = ["Date of Loss", "State", "City", "Accident Type",
                "Total Amount Billed ($)", "Medical Bills ($)",
                "Vehicles in Accident", "Fraud Flag", "Police Report on File",
                "Witness Count", "Risk Score", "Risk Level", "Risk Flags"]
    csv_cols = [c for c in csv_cols if c in df.columns]
    df[csv_cols].to_csv(f"{OUT_DIR}/nofault_scored.csv", index=False)
    print(f"  Scored CSV saved: {OUT_DIR}/nofault_scored.csv")

    rc = df["Risk Level"].value_counts()
    print(f"\n  Risk summary:")
    for level in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        if level in rc:
            print(f"    {level:<10}: {rc[level]:>4} claims")

    print("\n  Generating charts...")
    weekly = chart_weekly_volume(df)
    chart_arima_forecast(weekly)
    chart_seasonality(df)
    chart_fraud_volume(df)
    chart_top_cities(df)
    chart_risk_distribution(df)

    reserve_forecast(df)
    city_alerts(df)

    print("\n  Phase 2 complete.\n")

# =============================================================================
# PHASE 3 — REPORT & DELIVER
# =============================================================================

DARK_BLUE  = HexColor("#1F3864")
MED_BLUE   = HexColor("#2E75B6")
LIGHT_BLUE = HexColor("#BDD7EE")
PALE_BLUE  = HexColor("#EBF3FB")
WHITE      = HexColor("#FFFFFF")
DARK_GRAY  = HexColor("#404040")
MID_GRAY   = HexColor("#808080")
LIGHT_GRAY = HexColor("#F5F5F5")
PDF_RED    = HexColor("#C00000")
PDF_ORANGE = HexColor("#ED7D31")
PDF_GREEN  = HexColor("#70AD47")
YELLOW     = HexColor("#FFD966")

PDF_RISK_COLORS = {
    "LOW":      PDF_GREEN,
    "MEDIUM":   YELLOW,
    "HIGH":     PDF_ORANGE,
    "CRITICAL": PDF_RED,
}

def _load_scored():
    conn = sqlite3.connect(DB_FILE)
    try:
        df = pd.read_sql("SELECT * FROM claims_scored", conn)
        conn.close()
        return df
    except Exception:
        conn.close()
        return None

def _style(name, **kw):
    return ParagraphStyle(name, **kw)

def header_table(text, bg=None, fg=WHITE, size=12):
    if bg is None:
        bg = DARK_BLUE
    t = Table([[Paragraph(f"<b>{text}</b>",
                _style("h", fontName="Helvetica-Bold", fontSize=size,
                       textColor=fg, leading=size + 4))]],
              colWidths=[7 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), bg),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
    ]))
    return t

def kpi_row(kpis):
    cells = []
    for label, value, color in kpis:
        cell = Table([
            [Paragraph(str(value), _style("kv", fontName="Helvetica-Bold",
                        fontSize=20, textColor=color, alignment=TA_CENTER))],
            [Paragraph(label, _style("kl", fontName="Helvetica", fontSize=8,
                        textColor=DARK_GRAY, alignment=TA_CENTER))],
        ], colWidths=[1.55 * inch])
        cell.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), LIGHT_GRAY),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("BOX",           (0, 0), (-1, -1), 0.5, LIGHT_BLUE),
        ]))
        cells.append(cell)
    row = Table([cells], colWidths=[1.65 * inch] * len(kpis))
    row.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    return row

def build_pdf(df):
    today = datetime.date.today().strftime("%B %d, %Y")
    out   = os.path.join(OUT_DIR, "weekly_report.pdf")
    doc   = SimpleDocTemplate(out, pagesize=letter,
                              leftMargin=0.75*inch, rightMargin=0.75*inch,
                              topMargin=0.6*inch, bottomMargin=0.6*inch)
    story = []

    cover = Table([[
        Paragraph("No-Fault Claims Intelligence System",
                  _style("ct", fontName="Helvetica-Bold", fontSize=18,
                         textColor=WHITE, leading=22)),
    ]], colWidths=[7 * inch])
    cover.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), DARK_BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 20),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 16),
    ]))
    sub = Table([[
        Paragraph(f"Weekly Intelligence Briefing  ·  Generated {today}",
                  _style("cs", fontName="Helvetica", fontSize=10,
                         textColor=LIGHT_BLUE, leading=14)),
    ]], colWidths=[7 * inch])
    sub.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), DARK_BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
        ("LEFTPADDING",   (0, 0), (-1, -1), 16),
    ]))
    story += [cover, sub, Spacer(1, 12)]

    total    = len(df)
    critical = (df["Risk Level"] == "CRITICAL").sum()
    high     = (df["Risk Level"] == "HIGH").sum()
    fraud    = (df.get("Fraud Flag", df.iloc[:, 0].map(lambda x: "N")) == "Y").sum()
    fraud_pct = f"{fraud / total * 100:.1f}%" if total else "0%"

    story.append(header_table("Executive Summary — Key Performance Indicators"))
    story.append(Spacer(1, 6))
    story.append(kpi_row([
        ("Total claims",  f"{total:,}",  MED_BLUE),
        ("Critical risk", str(critical), PDF_RED),
        ("High risk",     str(high),     PDF_ORANGE),
        ("Fraud flagged", str(fraud),    PDF_RED),
        ("Fraud rate",    fraud_pct,     PDF_ORANGE),
    ]))
    story.append(Spacer(1, 14))

    story.append(header_table("Risk Level Breakdown", MED_BLUE))
    story.append(Spacer(1, 6))
    rc = df["Risk Level"].value_counts()
    risk_data = [[
        Paragraph("<b>Risk Level</b>",    _style("rh",  fontName="Helvetica-Bold", fontSize=10, textColor=WHITE)),
        Paragraph("<b>Claims</b>",        _style("rh2", fontName="Helvetica-Bold", fontSize=10, textColor=WHITE, alignment=TA_CENTER)),
        Paragraph("<b>% of Total</b>",    _style("rh3", fontName="Helvetica-Bold", fontSize=10, textColor=WHITE, alignment=TA_CENTER)),
        Paragraph("<b>Action Required</b>",_style("rh4",fontName="Helvetica-Bold", fontSize=10, textColor=WHITE)),
    ]]
    actions = {
        "CRITICAL": "Immediate SIU referral — suspend payment pending review",
        "HIGH":     "Supervisor review + request additional documentation",
        "MEDIUM":   "Enhanced review — flag for follow-up within 5 days",
        "LOW":      "Standard processing — routine examiner review",
    }
    for level in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        cnt = rc.get(level, 0)
        pct = f"{cnt / total * 100:.1f}%" if total else "0%"
        risk_data.append([
            Paragraph(level, _style(f"rl{level}", fontName="Helvetica-Bold",
                        fontSize=10, textColor=PDF_RISK_COLORS[level])),
            Paragraph(str(cnt), _style("rc", fontName="Helvetica", fontSize=10, alignment=TA_CENTER)),
            Paragraph(pct,      _style("rp", fontName="Helvetica", fontSize=10, alignment=TA_CENTER)),
            Paragraph(actions[level], _style("ra", fontName="Helvetica", fontSize=9, textColor=DARK_GRAY)),
        ])
    rt = Table(risk_data, colWidths=[1.1*inch, 0.9*inch, 0.9*inch, 4.1*inch])
    rt.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), MED_BLUE),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("BOX",           (0, 0), (-1, -1), 0.5, LIGHT_BLUE),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, HexColor("#DDDDDD")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(rt)
    story.append(Spacer(1, 14))

    chart_files = [
        ("chart_1_weekly_volume.png",    "Chart 1 — Weekly claim volume with 4-week moving average"),
        ("chart_2_ma_vs_arima.png",      "Chart 2 — Moving average vs ARIMA 8-week forecast"),
        ("chart_3_seasonality.png",      "Chart 3 — Seasonal claim distribution by month"),
        ("chart_4_fraud_vs_volume.png",  "Chart 4 — Claim volume vs fraud-flagged claims"),
        ("chart_5_top_cities.png",       "Chart 5 — Top 10 billing groups"),
        ("chart_6_risk_distribution.png","Chart 6 — Risk level distribution"),
    ]
    story.append(header_table("Charts & Visual Analysis", MED_BLUE))
    story.append(Spacer(1, 8))
    body_s = _style("body", fontName="Helvetica", fontSize=10,
                    textColor=DARK_GRAY, leading=15, alignment=TA_JUSTIFY)
    for fname, caption in chart_files:
        path = os.path.join(OUT_DIR, fname)
        if os.path.exists(path):
            story.append(Paragraph(f"<b>{caption}</b>", body_s))
            story.append(Spacer(1, 4))
            story.append(Image(path, width=7*inch, height=3*inch))
            story.append(Spacer(1, 12))

    story.append(PageBreak())
    story.append(header_table("Top Critical & High Risk Claims — Review Immediately", PDF_RED))
    story.append(Spacer(1, 6))
    top_risk = df[df["Risk Level"].isin(["CRITICAL", "HIGH"])].nlargest(15, "Risk Score")
    if len(top_risk) == 0:
        story.append(Paragraph("No critical or high-risk claims found.", body_s))
    else:
        show_cols = ["Date of Loss", "State", "Accident Type",
                     "Total Amount Billed ($)", "Risk Score", "Risk Level", "Risk Flags"]
        show_cols = [c for c in show_cols if c in top_risk.columns]
        header_row = [Paragraph(f"<b>{c}</b>", _style("th_s", fontName="Helvetica-Bold",
                                  fontSize=8, textColor=WHITE)) for c in show_cols]
        rows = [header_row]
        for _, row in top_risk.iterrows():
            cells = []
            for c in show_cols:
                val = row[c]
                if c == "Total Amount Billed ($)":
                    val = f"${float(val):,.0f}" if val else "$0"
                elif c == "Date of Loss" and hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                color = PDF_RISK_COLORS.get(str(row.get("Risk Level", "LOW")), DARK_GRAY)
                cells.append(Paragraph(str(val)[:60],
                             _style(f"td_{c}", fontName="Helvetica", fontSize=7.5,
                                    textColor=color if c == "Risk Level" else DARK_GRAY)))
            rows.append(cells)
        col_w = [w * inch for w in [0.9, 0.5, 0.9, 1.0, 0.7, 0.7, 2.3][:len(show_cols)]]
        ct = Table(rows, colWidths=col_w, repeatRows=1)
        ct.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), PDF_RED),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("BOX",           (0, 0), (-1, -1), 0.5, HexColor("#DDDDDD")),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, HexColor("#EEEEEE")),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(ct)

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MED_BLUE))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Generated by No-Fault Claims Intelligence System  ·  {today}  ·  "
        "For internal use only — not for distribution",
        _style("footer", fontName="Helvetica-Oblique", fontSize=8,
               textColor=MID_GRAY, alignment=TA_CENTER)))

    doc.build(story)
    print(f"  PDF report saved: {out}")

def update_excel(df):
    try:
        import openpyxl
    except ImportError:
        print("  Skipping Excel update — openpyxl not installed.")
        return
    xlsx_path = "excel/nofault_claims_toolkit.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"  Excel toolkit not found at '{xlsx_path}' — skipping update.")
        return
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    if "📋 Claim Log" in wb.sheetnames:
        ws = wb["📋 Claim Log"]
        risk_col = None
        for cell in ws[3]:
            if cell.value and "Risk" in str(cell.value):
                risk_col = cell.column
                break
        if risk_col:
            for i, (_, row) in enumerate(df.head(50).iterrows(), 4):
                ws.cell(row=i, column=risk_col, value=row.get("Risk Score", 0))
            print("  Excel toolkit updated with fresh risk scores.")
    wb.save(xlsx_path)

def phase3():
    print("\n" + "="*60)
    print("  PHASE 3 — REPORT & DELIVER")
    print("="*60)

    if not os.path.exists(DB_FILE):
        print("  ERROR: claims.db not found. Phases 1 and 2 must run first.")
        sys.exit(1)

    df = _load_scored()
    if df is None or len(df) == 0:
        print("  ERROR: No scored claims found. Phase 2 must run first.")
        sys.exit(1)

    print(f"\n  Loaded {len(df):,} scored claims")
    os.makedirs(OUT_DIR, exist_ok=True)

    print("\n  Building PDF report...")
    build_pdf(df)

    print("\n  Checking Excel toolkit...")
    update_excel(df)

    print("\n  Phase 3 complete.\n")

# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    start_all = time.time()
    print("\n" + "="*60)
    print("  ClaimsIQ — No-Fault Claims Intelligence System")
    print(f"  Started: {datetime.datetime.now():%Y-%m-%d %H:%M}")
    print("="*60)

    phase1()
    phase2()
    phase3()

    total = round(time.time() - start_all, 1)
    print("="*60)
    print(f"  Pipeline complete — {total}s total")
    print(f"  Outputs saved to: {OUT_DIR}/")
    print("="*60 + "\n")
