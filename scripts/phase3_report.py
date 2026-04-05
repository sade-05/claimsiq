# =============================================================================
# PHASE 3 — REPORT & DELIVER
# No-Fault Claims Intelligence System
# =============================================================================
# What this does:
#   1. Reads scored claims from the database
#   2. Generates a professional weekly PDF briefing report
#   3. Updates the Excel claims toolkit with fresh risk scores
# =============================================================================

import sqlite3
import os
import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable, Image, PageBreak
)
from reportlab.lib.colors import HexColor

DB_FILE  = "outputs/claims.db"
OUT_DIR  = "outputs"

# ── Colors ────────────────────────────────────────────────────────────────────
DARK_BLUE  = HexColor("#1F3864")
MED_BLUE   = HexColor("#2E75B6")
LIGHT_BLUE = HexColor("#BDD7EE")
PALE_BLUE  = HexColor("#EBF3FB")
WHITE      = HexColor("#FFFFFF")
DARK_GRAY  = HexColor("#404040")
MID_GRAY   = HexColor("#808080")
LIGHT_GRAY = HexColor("#F5F5F5")
RED        = HexColor("#C00000")
ORANGE     = HexColor("#ED7D31")
GREEN      = HexColor("#70AD47")
YELLOW     = HexColor("#FFD966")

RISK_COLORS = {
    "LOW":      GREEN,
    "MEDIUM":   YELLOW,
    "HIGH":     ORANGE,
    "CRITICAL": RED,
}

def load_scored():
    conn = sqlite3.connect(DB_FILE)
    try:
        import pandas as pd
        df = pd.read_sql("SELECT * FROM claims_scored", conn)
        conn.close()
        return df
    except Exception:
        conn.close()
        return None

def style(name, **kw):
    return ParagraphStyle(name, **kw)

def header_table(text, bg=DARK_BLUE, fg=WHITE, size=12):
    t = Table([[Paragraph(f"<b>{text}</b>",
                style("h", fontName="Helvetica-Bold", fontSize=size,
                      textColor=fg, leading=size + 4))]],
              colWidths=[7 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), bg),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
    ]))
    return t

def kpi_row(kpis):
    cells = []
    for label, value, color in kpis:
        cell = Table([
            [Paragraph(str(value), style("kv", fontName="Helvetica-Bold",
                        fontSize=20, textColor=color, alignment=TA_CENTER))],
            [Paragraph(label, style("kl", fontName="Helvetica", fontSize=8,
                        textColor=DARK_GRAY, alignment=TA_CENTER))],
        ], colWidths=[1.55 * inch])
        cell.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), LIGHT_GRAY),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("BOX",           (0, 0), (-1, -1), 0.5, LIGHT_BLUE),
        ]))
        cells.append(cell)
    row = Table([cells], colWidths=[1.65 * inch] * len(kpis))
    row.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    return row

def build_pdf(df):
    today   = datetime.date.today().strftime("%B %d, %Y")
    out     = os.path.join(OUT_DIR, "weekly_report.pdf")
    doc     = SimpleDocTemplate(out, pagesize=letter,
                                leftMargin=0.75*inch, rightMargin=0.75*inch,
                                topMargin=0.6*inch, bottomMargin=0.6*inch)
    story   = []

    # Cover header
    cover = Table([[
        Paragraph("No-Fault Claims Intelligence System",
                  style("ct", fontName="Helvetica-Bold", fontSize=18,
                        textColor=WHITE, leading=22)),
    ]], colWidths=[7 * inch])
    cover.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), DARK_BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 20),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 16),
    ]))
    sub = Table([[
        Paragraph(f"Weekly Intelligence Briefing  ·  Generated {today}",
                  style("cs", fontName="Helvetica", fontSize=10,
                        textColor=LIGHT_BLUE, leading=14)),
    ]], colWidths=[7 * inch])
    sub.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), DARK_BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
        ("LEFTPADDING",   (0, 0), (-1, -1), 16),
    ]))
    story += [cover, sub, Spacer(1, 12)]

    # KPIs
    total    = len(df)
    critical = (df["Risk Level"] == "CRITICAL").sum()
    high     = (df["Risk Level"] == "HIGH").sum()
    fraud    = (df.get("Fraud Flag", df.iloc[:, 0].map(lambda x: "N")) == "Y").sum()
    fraud_pct = f"{fraud / total * 100:.1f}%" if total else "0%"

    story.append(header_table("Executive Summary — Key Performance Indicators"))
    story.append(Spacer(1, 6))
    story.append(kpi_row([
        ("Total claims",         f"{total:,}",     MED_BLUE),
        ("Critical risk",        str(critical),    RED),
        ("High risk",            str(high),        ORANGE),
        ("Fraud flagged",        str(fraud),       RED),
        ("Fraud rate",           fraud_pct,        ORANGE),
    ]))
    story.append(Spacer(1, 14))

    # Risk breakdown table
    story.append(header_table("Risk Level Breakdown", MED_BLUE))
    story.append(Spacer(1, 6))
    rc = df["Risk Level"].value_counts()
    risk_data = [
        [Paragraph("<b>Risk Level</b>", style("rh", fontName="Helvetica-Bold",
                    fontSize=10, textColor=WHITE)),
         Paragraph("<b>Claims</b>", style("rh2", fontName="Helvetica-Bold",
                    fontSize=10, textColor=WHITE, alignment=TA_CENTER)),
         Paragraph("<b>% of Total</b>", style("rh3", fontName="Helvetica-Bold",
                    fontSize=10, textColor=WHITE, alignment=TA_CENTER)),
         Paragraph("<b>Action Required</b>", style("rh4", fontName="Helvetica-Bold",
                    fontSize=10, textColor=WHITE))],
    ]
    actions = {
        "CRITICAL": "Immediate SIU referral — suspend payment pending review",
        "HIGH":     "Supervisor review + request additional documentation",
        "MEDIUM":   "Enhanced review — flag for follow-up within 5 days",
        "LOW":      "Standard processing — routine examiner review",
    }
    for level in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        cnt = rc.get(level, 0)
        pct = f"{cnt / total * 100:.1f}%" if total else "0%"
        risk_data.append([
            Paragraph(level, style(f"rl{level}", fontName="Helvetica-Bold",
                        fontSize=10, textColor=RISK_COLORS[level])),
            Paragraph(str(cnt), style("rc", fontName="Helvetica",
                        fontSize=10, alignment=TA_CENTER)),
            Paragraph(pct, style("rp", fontName="Helvetica",
                        fontSize=10, alignment=TA_CENTER)),
            Paragraph(actions[level], style("ra", fontName="Helvetica",
                        fontSize=9, textColor=DARK_GRAY)),
        ])
    rt = Table(risk_data, colWidths=[1.1*inch, 0.9*inch, 0.9*inch, 4.1*inch])
    rt.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), MED_BLUE),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("BOX",           (0, 0), (-1, -1), 0.5, LIGHT_BLUE),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, HexColor("#DDDDDD")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(rt)
    story.append(Spacer(1, 14))

    # Charts
    chart_files = [
        ("chart_1_weekly_volume.png",  "Chart 1 — Weekly claim volume with 4-week moving average"),
        ("chart_2_ma_vs_arima.png",    "Chart 2 — Moving average vs ARIMA 8-week forecast"),
        ("chart_3_seasonality.png",    "Chart 3 — Seasonal claim distribution by month"),
        ("chart_4_fraud_vs_volume.png","Chart 4 — Claim volume vs fraud-flagged claims"),
        ("chart_5_billing_trends.png", "Chart 5 — Top 10 billing groups"),
        ("chart_8_risk_distribution.png","Chart 6 — Risk level distribution"),
    ]
    story.append(header_table("Charts & Visual Analysis", MED_BLUE))
    story.append(Spacer(1, 8))

    body_s = style("body", fontName="Helvetica", fontSize=10,
                   textColor=DARK_GRAY, leading=15, alignment=TA_JUSTIFY)

    for fname, caption in chart_files:
        path = os.path.join(OUT_DIR, fname)
        if os.path.exists(path):
            story.append(Paragraph(f"<b>{caption}</b>", body_s))
            story.append(Spacer(1, 4))
            story.append(Image(path, width=7*inch, height=3*inch))
            story.append(Spacer(1, 12))

    geo_charts = [
        ("chart_6_geo_claim_density.png", "Chart 7 — Claim density by state (choropleth)"),
        ("chart_7_geo_fraud_rate.png",    "Chart 8 — Fraud rate by state (choropleth)"),
    ]
    for fname, caption in geo_charts:
        path = os.path.join(OUT_DIR, fname)
        if os.path.exists(path):
            story.append(Paragraph(f"<b>{caption}</b>", body_s))
            story.append(Spacer(1, 4))
            story.append(Image(path, width=7*inch, height=3.5*inch))
            story.append(Spacer(1, 12))

    # Top critical claims table
    story.append(PageBreak())
    story.append(header_table("Top Critical & High Risk Claims — Review Immediately", RED))
    story.append(Spacer(1, 6))

    top_risk = df[df["Risk Level"].isin(["CRITICAL", "HIGH"])].nlargest(15, "Risk Score")
    if len(top_risk) == 0:
        story.append(Paragraph("No critical or high-risk claims found.", body_s))
    else:
        show_cols = ["Date of Loss", "State", "Accident Type",
                     "Total Amount Billed ($)", "Risk Score", "Risk Level", "Risk Flags"]
        show_cols = [c for c in show_cols if c in top_risk.columns]
        header = [Paragraph(f"<b>{c}</b>", style("th_s", fontName="Helvetica-Bold",
                              fontSize=8, textColor=WHITE))
                  for c in show_cols]
        rows = [header]
        for _, row in top_risk.iterrows():
            cells = []
            for c in show_cols:
                val = row[c]
                if c == "Total Amount Billed ($)":
                    val = f"${float(val):,.0f}" if val else "$0"
                elif c == "Date of Loss" and hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                color = RISK_COLORS.get(str(row.get("Risk Level", "LOW")), DARK_GRAY)
                cells.append(Paragraph(str(val)[:60],
                             style(f"td_{c}", fontName="Helvetica", fontSize=7.5,
                                   textColor=color if c == "Risk Level" else DARK_GRAY)))
            rows.append(cells)

        col_w = [0.9, 0.5, 0.9, 1.0, 0.7, 0.7, 2.3]
        col_w = [w * inch for w in col_w[:len(show_cols)]]
        ct = Table(rows, colWidths=col_w, repeatRows=1)
        ct.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), RED),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("BOX",           (0, 0), (-1, -1), 0.5, HexColor("#DDDDDD")),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, HexColor("#EEEEEE")),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(ct)

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MED_BLUE))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Generated by No-Fault Claims Intelligence System  ·  {today}  ·  "
        "For internal use only — not for distribution",
        style("footer", fontName="Helvetica-Oblique", fontSize=8,
              textColor=MID_GRAY, alignment=TA_CENTER)))

    doc.build(story)
    print(f"  PDF report saved: {out}")

def update_excel(df):
    try:
        import openpyxl
    except ImportError:
        print("  Skipping Excel update — openpyxl not installed.")
        return

    xlsx_path = "excel/nofault_claims_toolkit.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"  Excel toolkit not found at '{xlsx_path}' — skipping update.")
        print("  Place nofault_claims_toolkit.xlsx in the excel/ folder to enable this.")
        return

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)

    if "📋 Claim Log" in wb.sheetnames:
        ws = wb["📋 Claim Log"]
        risk_col = None
        for cell in ws[3]:
            if cell.value and "Risk" in str(cell.value):
                risk_col = cell.column
                break

        if risk_col:
            for i, (_, row) in enumerate(df.head(50).iterrows(), 4):
                ws.cell(row=i, column=risk_col, value=row.get("Risk Score", 0))
            print("  Excel toolkit updated with fresh risk scores.")

    wb.save(xlsx_path)

def run():
    print("\n" + "="*60)
    print("  PHASE 3 — REPORT & DELIVER")
    print("="*60)

    if not os.path.exists(DB_FILE):
        print("  ERROR: claims.db not found. Run phases 1 and 2 first.")
        return False

    df = load_scored()
    if df is None or len(df) == 0:
        print("  ERROR: No scored claims found. Run phase2_forecast.py first.")
        return False

    print(f"\n  Loaded {len(df):,} scored claims")
    os.makedirs(OUT_DIR, exist_ok=True)

    print("\n  Building PDF report...")
    build_pdf(df)

    print("\n  Checking Excel toolkit...")
    update_excel(df)

    print("\n  Phase 3 complete.\n")
    return True

if __name__ == "__main__":
    run()
